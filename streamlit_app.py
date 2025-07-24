import streamlit as st
import pandas as pd
import io
import msoffcrypto
from openpyxl import load_workbook

# 🔧 처리 함수
def process_excel_file(file_obj, password):
    decrypted = io.BytesIO()
    file = msoffcrypto.OfficeFile(file_obj)
    file.load_key(password=password)
    file.decrypt(decrypted)
    decrypted.seek(0)

    wb = load_workbook(filename=decrypted, data_only=True)
    output_sheets = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        data = list(ws.values)

        # ✅ 비어있지 않은 줄을 헤더로 자동 인식
        for i, row in enumerate(data):
            if row and any(cell is not None for cell in row):
                header = row
                content = data[i+1:]
                break
        else:
            continue  # 모든 줄이 비어 있음 → 다음 시트로

        df = pd.DataFrame(content, columns=header)
        df = df.fillna("").astype(str)

        if '예약의사' not in df or '예약시간' not in df:
            continue

        df['예약의사'] = df['예약의사'].str.replace(" 교수님", "", regex=False).strip()
        df = df.sort_values(by=['예약의사', '예약시간'])

        # 예약의사 기준 줄바꿈
        result = []
        current_doc = None
        for _, row in df.iterrows():
            if row['예약의사'] != current_doc:
                if current_doc is not None:
                    result.append([" "] * len(df.columns))
                current_doc = row['예약의사']
            result.append(row.tolist())

        final_df = pd.DataFrame(result, columns=df.columns)
        output_sheets[sheet] = final_df

    if not output_sheets:
        return None

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for name, sheet_df in output_sheets.items():
            sheet_df.to_excel(writer, sheet_name=name, index=False)
    output.seek(0)
    return output

# 🌐 Streamlit UI
st.set_page_config(page_title="엑셀 자동 처리기", layout="centered")
st.title("📄 암호화된 엑셀 자동 처리기")
uploaded = st.file_uploader("엑셀 파일 업로드", type=["xlsx"])
password = st.text_input("비밀번호 입력", type="password")

if uploaded and password:
    if st.button("처리하기"):
        try:
            result = process_excel_file(uploaded, password)
            if result:
                st.success("✅ 처리 완료! 아래에서 다운로드하세요.")
                st.download_button("📥 결과 다운로드", result, file_name="처리된_파일.xlsx")
            else:
                st.error("⚠ 유효한 데이터가 없습니다.")
        except Exception as e:
            st.error(f"❌ 에러 발생: {e}")
